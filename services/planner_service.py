import json
import os
import tempfile
import time
import traceback
from datetime import date, timedelta
from itertools import groupby

import tiktoken
import pandas as pd
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field

from prompts import get_course_planner_prompt
from utils.llm_utils import get_llm_from_env


def _count_tokens(text: str, model: str = "gpt-4o") -> int:
    """Estimate token count using tiktoken."""
    try:
        enc = tiktoken.encoding_for_model(model)
    except KeyError:
        enc = tiktoken.get_encoding("cl100k_base")
    return len(enc.encode(text))

INPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "data", "input")
_DEFAULT_OUTPUT_DIR = os.path.abspath(
    os.path.join(os.path.dirname(__file__), "..", "data", "output")
)
OUTPUT_DIR = os.environ.get("SMART_PLANNER_OUTPUT_DIR") or _DEFAULT_OUTPUT_DIR

TRAINER_LEAVE_FILE = "trainer_holidays_sep_to_dec_2026.xlsx"
PRIORITY_FILE = "priority_to_train_list.xlsx"
LOCATION_FILE = "Smart course planner Trainer per location.xlsx"
TRAINERS_JSON_FILE = "trainers.json"
COURSE_SCHEDULE_FILE = "course_schedule_days.xlsx"

# Planning period: September to December 2026
PLANNING_START = date(2026, 9, 1)
PLANNING_END = date(2026, 12, 31)


# --- Pydantic models for structured LLM output ---

class CourseSession(BaseModel):
    course_code: str = Field(description="Course code identifier")
    course_name: str = Field(description="Name of the course")
    session_number: int = Field(description="Session number (1, 2, 3, ...)")
    start_date: str = Field(description="Start date for the session (YYYY-MM-DD)")
    end_date: str = Field(description="End date for the session (YYYY-MM-DD)")
    trainer_name: str = Field(description="Assigned trainer full name")


class CoursePlan(BaseModel):
    sessions: list[CourseSession] = Field(
        description="List of all planned course sessions with dates and trainers"
    )


def _load_trainer_leave_dates() -> tuple[list[dict], str]:
    """Load trainer leave/holiday dates from trainer_holidays_sep_to_dec_2026.xlsx."""
    path = os.path.join(INPUT_DIR, TRAINER_LEAVE_FILE)
    df = pd.read_excel(path)

    trainers = []
    for _, row in df.iterrows():
        trainer_name = row.get("Trainer Name")
        if pd.isna(trainer_name):
            continue

        raw_leave_dates = row.get("Holiday / Leave Dates (2026)")
        leave_dates = []
        if not pd.isna(raw_leave_dates):
            for raw_date in str(raw_leave_dates).split(","):
                raw_date = raw_date.strip()
                # Format is "09 September" — append year 2026 for parsing
                parsed_date = pd.to_datetime(
                    f"{raw_date} 2026", format="%d %B %Y", errors="coerce"
                )
                if not pd.isna(parsed_date):
                    leave_dates.append(parsed_date.strftime("%Y-%m-%d"))

        trainers.append(
            {
                "trainer": str(trainer_name).strip(),
                "leave_dates": sorted(set(leave_dates)),
            }
        )

    period = f"{PLANNING_START} to {PLANNING_END}"
    return trainers, period


def _load_trainer_priority() -> pd.DataFrame:
    """Load trainer priority list. The file has header row in the data,
    so we parse accordingly."""
    path = os.path.join(INPUT_DIR, PRIORITY_FILE)
    df = pd.read_excel(path)
    # First data row is the sub-header: 1, 2, 3, 4, 5
    # Rename columns properly
    df.columns = [
        "course_code",
        "course_name",
        "priority_1",
        "priority_2",
        "priority_3",
        "priority_4",
        "priority_5",
    ]
    # Drop the header row (row 0 which has '1','2','3','4','5')
    df = df.iloc[1:].reset_index(drop=True)
    return df


def _priority_trainer_names(priority_df: pd.DataFrame) -> set[str]:
    priority_columns = [
        "priority_1",
        "priority_2",
        "priority_3",
        "priority_4",
        "priority_5",
    ]
    trainer_names = set()
    for column in priority_columns:
        trainer_names.update(
            str(name).strip()
            for name in priority_df[column].dropna().tolist()
            if str(name).strip()
        )
    return trainer_names


def _normalize_trainer_name(name: str) -> str:
    return " ".join(str(name).strip().lower().split())


def _trainer_name_aliases(name: str) -> set[str]:
    aliases = {_normalize_trainer_name(name)}
    if "," in str(name):
        last_name, first_names = [part.strip() for part in str(name).split(",", 1)]
        aliases.add(_normalize_trainer_name(f"{first_names} {last_name}"))
        last_name_parts = last_name.split()
        if last_name_parts:
            aliases.add(_normalize_trainer_name(f"{first_names} {last_name_parts[0]}"))
    return aliases


def _build_leave_lookup(trainers: list[dict]) -> dict[str, dict]:
    leave_lookup = {}
    for trainer in trainers:
        for alias in _trainer_name_aliases(trainer["trainer"]):
            leave_lookup[alias] = trainer
    return leave_lookup


def _load_trainer_locations() -> dict[str, dict]:
    """Load trainer→location mapping. Returns dict[normalized_alias → {location, country, display_name}]."""
    path = os.path.join(INPUT_DIR, LOCATION_FILE)
    try:
        df = pd.read_excel(path)
        result: dict[str, dict] = {}
        for _, row in df.iterrows():
            name = str(row.get("Trainer", "")).strip()
            if not name or name == "nan":
                continue
            info = {
                "location": str(row.get("Location", "")).strip(),
                "country": str(row.get("Country", "")).strip(),
                "display_name": name,
            }
            for alias in _trainer_name_aliases(name):
                result[alias] = info
        return result
    except Exception as e:
        print(f"[PLANNER] WARNING: Could not load location file: {e}")
        return {}


def _load_parttime_days_from_json() -> dict[str, set]:
    """Extract dates with '%' marker from trainers.json (part-time / half-day indicator).
    Returns dict[normalized_alias → set[date]]."""
    path = os.path.join(INPUT_DIR, TRAINERS_JSON_FILE)
    try:
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
        result: dict[str, set] = {}
        for trainer in data.get("trainers", []):
            tname = trainer["trainer"]
            pdays: set = set()
            for month_data in trainer.get("months", {}).values():
                for range_type in ("available_ranges", "on_leave_ranges"):
                    for r in month_data.get(range_type, []):
                        if r.get("marker") == "%":
                            rd_start = pd.to_datetime(r["start"], errors="coerce")
                            rd_end = pd.to_datetime(r["end"], errors="coerce")
                            if pd.isna(rd_start) or pd.isna(rd_end):
                                continue
                            cur = rd_start.date()
                            while cur <= rd_end.date():
                                pdays.add(cur)
                                cur += timedelta(days=1)
            for alias in _trainer_name_aliases(tname):
                result[alias] = pdays
        return result
    except Exception as e:
        print(f"[PLANNER] WARNING: Could not load parttime days from trainers.json: {e}")
        return {}


def _get_planning_period_from_json() -> tuple:
    """Parse the planning period string from trainers.json. Returns (start_date, end_date) or (None, None)."""
    path = os.path.join(INPUT_DIR, TRAINERS_JSON_FILE)
    try:
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
        period_str = data.get("period", "")
        if " to " in period_str:
            parts = period_str.split(" to ")
            start = pd.to_datetime(parts[0].strip(), errors="coerce")
            end = pd.to_datetime(parts[1].strip(), errors="coerce")
            if not pd.isna(start) and not pd.isna(end):
                return start.date(), end.date()
    except Exception as e:
        print(f"[PLANNER] WARNING: Could not parse planning period from trainers.json: {e}")
    return None, None


def _parse_classrooms(classroom_raw: str, session_start: date, session_end: date) -> list:
    """Parse classroom assignment string into list of (lab_name, start_date, end_date).
    Handles multi-lab strings like 'QS003 wk1+2\\nQS147 wk3'."""
    lines = [ln.strip() for ln in classroom_raw.strip().split("\n") if ln.strip()]
    if len(lines) <= 1:
        lab = classroom_raw.strip().split()[0] if classroom_raw.strip().split() else classroom_raw
        return [(lab, session_start, session_end)]

    result = []
    for line in lines:
        parts = line.split()
        if not parts:
            continue
        lab = parts[0]
        wk_part = next((p[2:] for p in parts[1:] if p.lower().startswith("wk")), None)
        if wk_part:
            week_nums = []
            for seg in wk_part.replace("+", ",").split(","):
                try:
                    week_nums.append(int(seg.strip()))
                except ValueError:
                    pass
            if week_nums:
                min_wk = min(week_nums) - 1
                max_wk = max(week_nums) - 1
                lab_start = session_start + timedelta(weeks=min_wk)
                lab_end = min(session_start + timedelta(weeks=max_wk + 1) - timedelta(days=1), session_end)
                result.append((lab, lab_start, lab_end))
                continue
        result.append((lab, session_start, session_end))
    return result if result else [(classroom_raw.split()[0] if classroom_raw.split() else classroom_raw, session_start, session_end)]


def _write_trainer_availability_sheet(
    output_path: str,
    trainers: list[dict],
    sessions: list[dict],
    priority_trainer_names: set[str],
) -> None:
    """Append a 'Trainer Availability' sheet to the plan workbook.

    Layout:
      Row 1 : month labels (merged across that month's columns)
      Row 2 : day numbers
      Row 3+ : trainer rows, grouped by location
    Colors:
      light blue   – teaching (course code shown; merged across consecutive days)
      light red    – on leave
      light yellow – part-time / half-day (% marker from source data)
      light green  – weekend
    A legend block is appended below the trainer rows.
    """
    # --- Colour palette ---
    LIGHT_BLUE   = PatternFill(start_color="FFADD8E6", end_color="FFADD8E6", fill_type="solid")
    LIGHT_RED    = PatternFill(start_color="FFFFB6C1", end_color="FFFFB6C1", fill_type="solid")
    WEEKEND_FILL = PatternFill(start_color="FFE8F5E9", end_color="FFE8F5E9", fill_type="solid")
    PARTTIME_FILL= PatternFill(start_color="FFFFF59D", end_color="FFFFF59D", fill_type="solid")
    HEADER_FILL  = PatternFill(start_color="FF1F4E78", end_color="FF1F4E78", fill_type="solid")
    MONTH_FILL   = PatternFill(start_color="FF2E75B6", end_color="FF2E75B6", fill_type="solid")
    DAYNUM_FILL  = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    LOC_FILL_MAP = {
        "PHC":    PatternFill(start_color="FFD6E4BC", end_color="FFD6E4BC", fill_type="solid"),
        "SLC":    PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid"),
        "CTC":    PatternFill(start_color="FFFCE4D6", end_color="FFFCE4D6", fill_type="solid"),
        "Others": PatternFill(start_color="FFF2EFEA", end_color="FFF2EFEA", fill_type="solid"),
    }
    DEFAULT_LOC_FILL = PatternFill(start_color="FFF2F2F2", end_color="FFF2F2F2", fill_type="solid")

    HEADER_FONT     = Font(bold=True, color="FFFFFFFF")
    SESSION_FONT    = Font(size=7, bold=True)
    SESSION_ALIGN   = Alignment(horizontal="center", vertical="center", wrap_text=False)
    LOC_HEADER_FONT = Font(bold=True, size=10)

    # --- Extra data ---
    location_lookup = _load_trainer_locations()
    parttime_lookup = _load_parttime_days_from_json()
    plan_start, plan_end = _get_planning_period_from_json()

    # --- Date range (planning quarter, not full year) ---
    if plan_start and plan_end:
        start_d, end_d = plan_start, plan_end
    else:
        all_dates: list[date] = []
        for s in sessions:
            for key in ("start_date", "end_date"):
                parsed = pd.to_datetime(s.get(key), errors="coerce")
                if not pd.isna(parsed):
                    all_dates.append(parsed.date())
        if not all_dates:
            print("[PLANNER] Skipping trainer availability sheet: no dates available")
            return
        start_d = date(min(all_dates).year, min(all_dates).month, 1)
        end_d   = max(all_dates)

    date_list: list[date] = []
    cur = start_d
    while cur <= end_d:
        date_list.append(cur)
        cur += timedelta(days=1)

    date_to_col: dict[date, int] = {dt: i + 2 for i, dt in enumerate(date_list)}
    total_cols = len(date_list) + 1  # col A + date columns

    # --- Trainer sessions lookup (alias-aware) ---
    trainer_sessions: dict[str, list[dict]] = {}
    for s in sessions:
        norm = _normalize_trainer_name(s.get("trainer_name", ""))
        if not norm:
            continue
        sd = pd.to_datetime(s.get("start_date"), errors="coerce")
        ed = pd.to_datetime(s.get("end_date"), errors="coerce")
        if pd.isna(sd) or pd.isna(ed):
            continue
        trainer_sessions.setdefault(norm, []).append({
            "start":          sd.date(),
            "end":            ed.date(),
            "course_code":    str(s.get("course_code",    "") or "").strip(),
            "course_name":    str(s.get("course_name",    "") or "").strip(),
            "session_number": s.get("session_number"),
        })

    leave_lookup = _build_leave_lookup(trainers)

    # --- Deduplicated trainer list (fix: alias-aware dedup prevents "Last, First" duplicates) ---
    trainer_names: list[str] = []
    seen: set[str] = set()

    def _add_trainer(name: str) -> None:
        cleaned = str(name).strip()
        if not cleaned:
            return
        aliases = _trainer_name_aliases(cleaned)
        if aliases & seen:   # any alias already seen → skip
            return
        seen.update(aliases)
        trainer_names.append(cleaned)

    for trainer in trainers:
        _add_trainer(trainer["trainer"])
    for name in sorted(priority_trainer_names):
        _add_trainer(name)
    for s in sessions:
        _add_trainer(s.get("trainer_name", ""))

    # --- Group trainers by location ---
    location_groups: dict[str, list[str]] = {}
    others: list[str] = []
    for name in trainer_names:
        info = None
        for alias in _trainer_name_aliases(name):
            if alias in location_lookup:
                info = location_lookup[alias]
                break
        if info:
            loc = info["location"]
            location_groups.setdefault(loc, []).append(name)
        else:
            others.append(name)
    if others:
        location_groups["Others"] = others

    # --- Create / replace sheet ---
    wb = load_workbook(output_path)
    if "Trainer Availability" in wb.sheetnames:
        del wb["Trainer Availability"]
    ws = wb.create_sheet("Trainer Availability")

    # --- Row 1: "Trainer" label (spans rows 1-2) + month headers ---
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    hc = ws.cell(row=1, column=1, value="Trainer")
    hc.fill = HEADER_FILL
    hc.font = HEADER_FONT
    hc.alignment = Alignment(horizontal="center", vertical="center")

    for (yr, mo), grp in groupby(date_list, key=lambda d: (d.year, d.month)):
        month_dates = list(grp)
        col_s = date_to_col[month_dates[0]]
        col_e = date_to_col[month_dates[-1]]
        if col_s < col_e:
            ws.merge_cells(start_row=1, start_column=col_s, end_row=1, end_column=col_e)
        mc = ws.cell(row=1, column=col_s, value=month_dates[0].strftime("%B %Y"))
        mc.fill = MONTH_FILL
        mc.font = HEADER_FONT
        mc.alignment = Alignment(horizontal="center", vertical="center")

    # --- Row 2: day numbers ---
    for dt in date_list:
        col = date_to_col[dt]
        dc = ws.cell(row=2, column=col, value=dt.day)
        dc.fill = DAYNUM_FILL
        dc.font = Font(bold=True, color="FFFFFFFF", size=7)
        dc.alignment = Alignment(horizontal="center", vertical="center")

    # --- Trainer rows, grouped by location ---
    current_row = 3
    trainer_row_indices: list[int] = []
    course_cells: dict[tuple, str] = {}  # (row, col) → course_code value

    for loc, names in location_groups.items():
        # Location group header
        loc_fill = LOC_FILL_MAP.get(loc, DEFAULT_LOC_FILL)
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=total_cols)
        lc = ws.cell(row=current_row, column=1, value=f"  {loc}")
        lc.fill = loc_fill
        lc.font = LOC_HEADER_FONT
        lc.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[current_row].height = 18
        current_row += 1

        for trainer_name in names:
            # Resolve leave dates via all aliases
            leave_set: set[date] = set()
            for alias in _trainer_name_aliases(trainer_name):
                record = leave_lookup.get(alias)
                if not record:
                    continue
                for raw in record.get("leave_dates", []):
                    parsed = pd.to_datetime(raw, errors="coerce")
                    if not pd.isna(parsed):
                        leave_set.add(parsed.date())

            # Resolve part-time days
            parttime_set: set[date] = set()
            for alias in _trainer_name_aliases(trainer_name):
                parttime_set.update(parttime_lookup.get(alias, set()))

            # Resolve session ranges
            ranges: list[dict] = []
            for alias in _trainer_name_aliases(trainer_name):
                ranges.extend(trainer_sessions.get(alias, []))

            ws.cell(row=current_row, column=1, value=trainer_name)

            for dt in date_list:
                col = date_to_col[dt]
                cell = ws.cell(row=current_row, column=col)
                matching = [r for r in ranges if r["start"] <= dt <= r["end"]]

                if matching:
                    cell.fill = LIGHT_BLUE
                    codes = [r["course_code"] for r in matching if r["course_code"]]
                    if codes:
                        val = " / ".join(dict.fromkeys(codes))
                        cell.value = val
                        cell.font  = SESSION_FONT
                        cell.alignment = SESSION_ALIGN
                        course_cells[(current_row, col)] = val
                    # Hover comment
                    comment_lines = []
                    for r in matching:
                        suffix = f" — Session {r['session_number']}" if r["session_number"] else ""
                        comment_lines.append(
                            f"{r['course_code'] or '?'}: {r['course_name'] or '(unknown)'}{suffix}"
                        )
                    if comment_lines:
                        cmt = Comment("\n".join(comment_lines), "Planner")
                        cmt.width  = 260
                        cmt.height = 20 + 14 * len(comment_lines)
                        cell.comment = cmt

                elif dt in leave_set:
                    cell.fill = LIGHT_RED
                    if dt in parttime_set:
                        cell.value     = "%"
                        cell.font      = Font(size=7, bold=True, color="FF8B0000")
                        cell.alignment = SESSION_ALIGN

                elif dt in parttime_set:
                    cell.fill      = PARTTIME_FILL
                    cell.value     = "%"
                    cell.font      = Font(size=7, color="FF856404")
                    cell.alignment = SESSION_ALIGN

                elif dt.weekday() >= 5:
                    cell.fill = WEEKEND_FILL

            trainer_row_indices.append(current_row)
            current_row += 1

    # --- Merge consecutive course-code cells within each trainer row ---
    for row_idx in trainer_row_indices:
        col = 2
        max_col = len(date_list) + 1
        while col <= max_col:
            pos = (row_idx, col)
            if pos in course_cells:
                val       = course_cells[pos]
                start_col = col
                while col + 1 <= max_col and course_cells.get((row_idx, col + 1)) == val:
                    col += 1
                end_col = col
                if end_col > start_col:
                    ws.merge_cells(start_row=row_idx, start_column=start_col,
                                   end_row=row_idx, end_column=end_col)
                    mc = ws.cell(row=row_idx, column=start_col)
                    mc.value     = val
                    mc.fill      = LIGHT_BLUE
                    mc.font      = SESSION_FONT
                    mc.alignment = SESSION_ALIGN
            col += 1

    # --- Legend ---
    legend_row = current_row + 1
    ws.cell(row=legend_row, column=1, value="Legend").font = Font(bold=True, size=10)
    legend_row += 1
    legend_items = [
        (LIGHT_BLUE,    "Teaching / Course assigned"),
        (LIGHT_RED,     "On leave"),
        (PARTTIME_FILL, "Part-time / half-day available (% marker)"),
        (WEEKEND_FILL,  "Weekend"),
    ]
    thin = Side(style="thin", color="FF999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for fill, label in legend_items:
        color_cell = ws.cell(row=legend_row, column=1, value=" ")
        color_cell.fill   = fill
        color_cell.border = border
        label_cell = ws.cell(row=legend_row, column=2, value=label)
        label_cell.font = Font(size=9)
        legend_row += 1

    # --- Column / row sizing ---
    ws.column_dimensions["A"].width = 28
    for col_idx in range(2, len(date_list) + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 4
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 18
    ws.freeze_panes = "B3"

    wb.save(output_path)


def _write_lab_availability_sheet(output_path: str, sessions: list[dict]) -> None:
    """Append a 'Lab Availability' sheet showing which labs/rooms are used per day,
    grouped by trainer location (PHC / SLC / CTC / Others)."""
    schedule_path = os.path.join(INPUT_DIR, COURSE_SCHEDULE_FILE)
    try:
        schedule_df = pd.read_excel(schedule_path)
    except Exception as e:
        print(f"[PLANNER] WARNING: Could not load course schedule for lab sheet: {e}")
        return

    # Build course → first classroom raw string
    classroom_lookup: dict[str, str] = {}
    for _, row in schedule_df.iterrows():
        code = str(row.get("Course Code", "")).strip()
        raw  = str(row.get("Preferred Classroom", "")).strip()
        if code and raw and raw.lower() != "nan" and code not in classroom_lookup:
            classroom_lookup[code] = raw

    location_lookup = _load_trainer_locations()

    plan_start, plan_end = _get_planning_period_from_json()
    if not plan_start or not plan_end:
        if not sessions:
            return
        sd_all = [pd.to_datetime(s.get("start_date"), errors="coerce").date()
                  for s in sessions]
        ed_all = [pd.to_datetime(s.get("end_date"),   errors="coerce").date()
                  for s in sessions]
        sd_all = [d for d in sd_all if d]
        ed_all = [d for d in ed_all if d]
        if not sd_all:
            return
        plan_start, plan_end = min(sd_all), max(ed_all)

    date_list: list[date] = []
    cur = plan_start
    while cur <= plan_end:
        date_list.append(cur)
        cur += timedelta(days=1)
    date_to_col = {dt: i + 2 for i, dt in enumerate(date_list)}
    total_cols = len(date_list) + 1

    # Build lab assignments and location map
    lab_assignments: dict[str, list[dict]] = {}
    lab_locations:   dict[str, str]        = {}

    for s in sessions:
        course_code  = str(s.get("course_code", "")).strip()
        trainer_name = str(s.get("trainer_name", "")).strip()
        sd = pd.to_datetime(s.get("start_date"), errors="coerce")
        ed = pd.to_datetime(s.get("end_date"),   errors="coerce")
        if pd.isna(sd) or pd.isna(ed):
            continue

        classroom_raw = classroom_lookup.get(course_code, "")
        if not classroom_raw or classroom_raw.lower() == "nan":
            classroom_raw = course_code  # fallback: use course code as lab identifier

        trainer_loc = "Others"
        for alias in _trainer_name_aliases(trainer_name):
            if alias in location_lookup:
                trainer_loc = location_lookup[alias]["location"]
                break

        for lab, lab_start, lab_end in _parse_classrooms(classroom_raw, sd.date(), ed.date()):
            lab_locations.setdefault(lab, trainer_loc)
            lab_assignments.setdefault(lab, []).append({
                "start":       lab_start,
                "end":         lab_end,
                "course_code": course_code,
                "course_name": str(s.get("course_name", "")).strip(),
                "trainer":     trainer_name,
            })

    if not lab_assignments:
        print("[PLANNER] No lab data to write – skipping Lab Availability sheet")
        return

    # Group labs by location, ensuring all location groups from trainer data are present
    loc_labs: dict[str, list[str]] = {}
    # Pre-populate all known location groups (same as trainer sheet) in order
    all_locations = list(dict.fromkeys(
        info["location"] for info in location_lookup.values() if info.get("location")
    ))
    for loc in all_locations:
        loc_labs.setdefault(loc, [])
    loc_labs.setdefault("Others", [])
    # Assign labs to their location groups
    for lab in sorted(lab_locations):
        loc = lab_locations[lab]
        loc_labs.setdefault(loc, []).append(lab)

    # --- Colours (same palette as trainer sheet) ---
    COURSE_FILL  = PatternFill(start_color="FFADD8E6", end_color="FFADD8E6", fill_type="solid")
    WEEKEND_FILL = PatternFill(start_color="FFE8F5E9", end_color="FFE8F5E9", fill_type="solid")
    HEADER_FILL  = PatternFill(start_color="FF1F4E78", end_color="FF1F4E78", fill_type="solid")
    MONTH_FILL   = PatternFill(start_color="FF2E75B6", end_color="FF2E75B6", fill_type="solid")
    DAYNUM_FILL  = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    LOC_FILL_MAP = {
        "PHC":    PatternFill(start_color="FFD6E4BC", end_color="FFD6E4BC", fill_type="solid"),
        "SLC":    PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid"),
        "CTC":    PatternFill(start_color="FFFCE4D6", end_color="FFFCE4D6", fill_type="solid"),
        "Others": PatternFill(start_color="FFF2EFEA", end_color="FFF2EFEA", fill_type="solid"),
    }
    DEFAULT_LOC_FILL = PatternFill(start_color="FFF2F2F2", end_color="FFF2F2F2", fill_type="solid")
    HEADER_FONT   = Font(bold=True, color="FFFFFFFF")
    SESSION_FONT  = Font(size=7, bold=True)
    SESSION_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=False)

    wb = load_workbook(output_path)
    if "Lab Availability" in wb.sheetnames:
        del wb["Lab Availability"]
    ws = wb.create_sheet("Lab Availability")

    # Row 1+2 header
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    hc = ws.cell(row=1, column=1, value="Lab / Room")
    hc.fill = HEADER_FILL
    hc.font = HEADER_FONT
    hc.alignment = Alignment(horizontal="center", vertical="center")

    for (yr, mo), grp in groupby(date_list, key=lambda d: (d.year, d.month)):
        month_dates = list(grp)
        col_s = date_to_col[month_dates[0]]
        col_e = date_to_col[month_dates[-1]]
        if col_s < col_e:
            ws.merge_cells(start_row=1, start_column=col_s, end_row=1, end_column=col_e)
        mc = ws.cell(row=1, column=col_s, value=month_dates[0].strftime("%B %Y"))
        mc.fill = MONTH_FILL
        mc.font = HEADER_FONT
        mc.alignment = Alignment(horizontal="center", vertical="center")

    for dt in date_list:
        col = date_to_col[dt]
        dc = ws.cell(row=2, column=col, value=dt.day)
        dc.fill = DAYNUM_FILL
        dc.font = Font(bold=True, color="FFFFFFFF", size=7)
        dc.alignment = Alignment(horizontal="center", vertical="center")

    current_row = 3
    lab_row_indices: list[int] = []
    course_cells: dict[tuple, str] = {}

    for loc, labs in loc_labs.items():
        loc_fill = LOC_FILL_MAP.get(loc, DEFAULT_LOC_FILL)
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=total_cols)
        lc = ws.cell(row=current_row, column=1, value=f"  {loc}")
        lc.fill = loc_fill
        lc.font = Font(bold=True, size=10)
        lc.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[current_row].height = 18
        current_row += 1

        for lab in labs:
            ws.cell(row=current_row, column=1, value=lab)
            lab_info = lab_assignments.get(lab, [])

            for dt in date_list:
                col  = date_to_col[dt]
                cell = ws.cell(row=current_row, column=col)
                matching = [r for r in lab_info if r["start"] <= dt <= r["end"]]
                if matching:
                    cell.fill = COURSE_FILL
                    codes = [r["course_code"] for r in matching if r["course_code"]]
                    if codes:
                        val = " / ".join(dict.fromkeys(codes))
                        cell.value     = val
                        cell.font      = SESSION_FONT
                        cell.alignment = SESSION_ALIGN
                        course_cells[(current_row, col)] = val
                    comment_lines = [
                        f"{r['course_code']}: {r['course_name']} ({r['trainer']})"
                        for r in matching
                    ]
                    if comment_lines:
                        cmt = Comment("\n".join(comment_lines), "Planner")
                        cmt.width  = 300
                        cmt.height = 20 + 14 * len(comment_lines)
                        cell.comment = cmt
                elif dt.weekday() >= 5:
                    cell.fill = WEEKEND_FILL

            lab_row_indices.append(current_row)
            current_row += 1

    # Merge consecutive same-course cells
    for row_idx in lab_row_indices:
        col = 2
        max_col = len(date_list) + 1
        while col <= max_col:
            pos = (row_idx, col)
            if pos in course_cells:
                val       = course_cells[pos]
                start_col = col
                while col + 1 <= max_col and course_cells.get((row_idx, col + 1)) == val:
                    col += 1
                end_col = col
                if end_col > start_col:
                    ws.merge_cells(start_row=row_idx, start_column=start_col,
                                   end_row=row_idx, end_column=end_col)
                    mc = ws.cell(row=row_idx, column=start_col)
                    mc.value     = val
                    mc.fill      = COURSE_FILL
                    mc.font      = SESSION_FONT
                    mc.alignment = SESSION_ALIGN
            col += 1

    ws.column_dimensions["A"].width = 22
    for col_idx in range(2, len(date_list) + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 4
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 18
    ws.freeze_panes = "B3"

    wb.save(output_path)


def _write_course_overview_sheet(
    output_path: str, courses: list[dict], sessions: list[dict]
) -> None:
    """Append a 'Course Overview' sheet with demand data and missed courses.

    Columns:
      Course Code | Course Name | Duration (Days) | Total Students |
      Sessions Needed | Sessions Planned | Sessions Missed | Missed Due to Timeline
    """

    HEADER_FILL = PatternFill(start_color="FF1F4E78", end_color="FF1F4E78", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFFFF")
    MISSED_FILL = PatternFill(start_color="FFFFCCCC", end_color="FFFFCCCC", fill_type="solid")
    ALIGN_CENTER = Alignment(horizontal="center", vertical="center")

    # Count sessions actually planned per course code
    planned_counts: dict[str, int] = {}
    for s in sessions:
        code = str(s.get("course_code", "")).strip()
        if code:
            planned_counts[code] = planned_counts.get(code, 0) + 1

    # Build overview rows
    overview_rows = []
    for c in courses:
        code = str(c.get("course_code", "")).strip()
        name = str(c.get("course_title", "")).strip()
        duration = c.get("no_of_days")
        students = c.get("students_interested")
        needed = c.get("sessions_needed")
        planned = planned_counts.get(code, 0)
        missed = max(0, (needed or 0) - planned) if needed else 0
        missed_flag = "Yes" if missed > 0 else "No"

        overview_rows.append({
            "course_code": code,
            "course_name": name,
            "duration_days": duration,
            "total_students": students,
            "sessions_needed": needed,
            "sessions_planned": planned,
            "sessions_missed": missed,
            "missed_due_to_timeline": missed_flag,
        })

    wb = load_workbook(output_path)
    if "Course Overview" in wb.sheetnames:
        del wb["Course Overview"]
    ws = wb.create_sheet("Course Overview")

    headers = [
        "Course Code", "Course Name", "Duration (Days)", "Total Students",
        "Sessions Needed", "Sessions Planned", "Sessions Missed",
        "Missed Due to Timeline",
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = ALIGN_CENTER

    for row_idx, row_data in enumerate(overview_rows, start=2):
        ws.cell(row=row_idx, column=1, value=row_data["course_code"])
        ws.cell(row=row_idx, column=2, value=row_data["course_name"])
        ws.cell(row=row_idx, column=3, value=row_data["duration_days"]).alignment = ALIGN_CENTER
        ws.cell(row=row_idx, column=4, value=row_data["total_students"]).alignment = ALIGN_CENTER
        ws.cell(row=row_idx, column=5, value=row_data["sessions_needed"]).alignment = ALIGN_CENTER
        ws.cell(row=row_idx, column=6, value=row_data["sessions_planned"]).alignment = ALIGN_CENTER
        missed_cell = ws.cell(row=row_idx, column=7, value=row_data["sessions_missed"])
        missed_cell.alignment = ALIGN_CENTER
        if row_data["sessions_missed"] > 0:
            missed_cell.fill = MISSED_FILL
        flag_cell = ws.cell(row=row_idx, column=8, value=row_data["missed_due_to_timeline"])
        flag_cell.alignment = ALIGN_CENTER
        if row_data["missed_due_to_timeline"] == "Yes":
            flag_cell.fill = MISSED_FILL

    # Summary row
    total_row = len(overview_rows) + 3
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=4, value=sum(
        r["total_students"] or 0 for r in overview_rows
    )).font = Font(bold=True)
    ws.cell(row=total_row, column=5, value=sum(
        r["sessions_needed"] or 0 for r in overview_rows
    )).font = Font(bold=True)
    ws.cell(row=total_row, column=6, value=sum(
        r["sessions_planned"] for r in overview_rows
    )).font = Font(bold=True)
    ws.cell(row=total_row, column=7, value=sum(
        r["sessions_missed"] for r in overview_rows
    )).font = Font(bold=True)

    # Auto-size columns
    col_widths = [14, 40, 16, 14, 16, 16, 16, 22]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(output_path)


def _write_utilization_metrics_sheet(
    output_path: str,
    trainers: list[dict],
    sessions: list[dict],
    priority_trainer_names: set[str],
) -> None:
    """Append a 'Utilization Metrics' sheet with:
    - Trainer utilization: teaching days vs total available days per quarter (%)
    - Lab/Room utilization: booked days vs total available days (%)
    """
    HEADER_FILL = PatternFill(start_color="FF1F4E78", end_color="FF1F4E78", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFFFF")
    SECTION_FONT = Font(bold=True, size=11)
    ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
    PCT_FORMAT = "0.0%"

    # --- Date range ---
    plan_start, plan_end = _get_planning_period_from_json()
    if not plan_start or not plan_end:
        plan_start, plan_end = PLANNING_START, PLANNING_END

    # Count total working days (Mon-Fri) in the planning period
    all_working_days: list[date] = []
    cur = plan_start
    while cur <= plan_end:
        if cur.weekday() < 5:
            all_working_days.append(cur)
        cur += timedelta(days=1)
    total_working_days = len(all_working_days)

    # --- Trainer Utilization ---
    leave_lookup = _build_leave_lookup(trainers)
    location_lookup = _load_trainer_locations()

    # Deduplicated trainer list
    trainer_names: list[str] = []
    seen: set[str] = set()
    for trainer in trainers:
        name = str(trainer["trainer"]).strip()
        aliases = _trainer_name_aliases(name)
        if not (aliases & seen):
            seen.update(aliases)
            trainer_names.append(name)
    for name in sorted(priority_trainer_names):
        aliases = _trainer_name_aliases(name)
        if not (aliases & seen):
            seen.update(aliases)
            trainer_names.append(name)
    for s in sessions:
        name = str(s.get("trainer_name", "")).strip()
        if name:
            aliases = _trainer_name_aliases(name)
            if not (aliases & seen):
                seen.update(aliases)
                trainer_names.append(name)

    # Build teaching days per trainer
    trainer_teaching_days: dict[str, set] = {}
    for s in sessions:
        norm = _normalize_trainer_name(s.get("trainer_name", ""))
        sd = pd.to_datetime(s.get("start_date"), errors="coerce")
        ed = pd.to_datetime(s.get("end_date"), errors="coerce")
        if pd.isna(sd) or pd.isna(ed):
            continue
        days_set = trainer_teaching_days.setdefault(norm, set())
        cur_d = sd.date()
        while cur_d <= ed.date():
            if cur_d.weekday() < 5:
                days_set.add(cur_d)
            cur_d += timedelta(days=1)

    trainer_metrics = []
    for name in trainer_names:
        # Leave days
        leave_set: set[date] = set()
        for alias in _trainer_name_aliases(name):
            record = leave_lookup.get(alias)
            if record:
                for raw in record.get("leave_dates", []):
                    parsed = pd.to_datetime(raw, errors="coerce")
                    if not pd.isna(parsed) and parsed.date().weekday() < 5:
                        leave_set.add(parsed.date())

        available_days = total_working_days - len(leave_set)

        # Teaching days
        teaching_days = 0
        for alias in _trainer_name_aliases(name):
            teaching_days += len(trainer_teaching_days.get(alias, set()))

        pct = teaching_days / available_days if available_days > 0 else 0

        # Location
        loc = "Others"
        for alias in _trainer_name_aliases(name):
            if alias in location_lookup:
                loc = location_lookup[alias]["location"]
                break

        trainer_metrics.append({
            "trainer": name,
            "location": loc,
            "total_working_days": total_working_days,
            "leave_days": len(leave_set),
            "available_days": available_days,
            "teaching_days": teaching_days,
            "utilization_pct": pct,
        })

    # --- Lab/Room Utilization ---
    schedule_path = os.path.join(INPUT_DIR, COURSE_SCHEDULE_FILE)
    classroom_lookup: dict[str, str] = {}
    try:
        schedule_df = pd.read_excel(schedule_path)
        for _, row in schedule_df.iterrows():
            code = str(row.get("Course Code", "")).strip()
            raw = str(row.get("Preferred Classroom", "")).strip()
            if code and raw and raw.lower() != "nan" and code not in classroom_lookup:
                classroom_lookup[code] = raw
    except Exception:
        pass

    lab_booked_days: dict[str, set] = {}
    for s in sessions:
        course_code = str(s.get("course_code", "")).strip()
        sd = pd.to_datetime(s.get("start_date"), errors="coerce")
        ed = pd.to_datetime(s.get("end_date"), errors="coerce")
        if pd.isna(sd) or pd.isna(ed):
            continue

        classroom_raw = classroom_lookup.get(course_code, "")
        if not classroom_raw or classroom_raw.lower() == "nan":
            classroom_raw = course_code

        for lab, lab_start, lab_end in _parse_classrooms(classroom_raw, sd.date(), ed.date()):
            days_set = lab_booked_days.setdefault(lab, set())
            cur_d = lab_start
            while cur_d <= lab_end:
                if cur_d.weekday() < 5:
                    days_set.add(cur_d)
                cur_d += timedelta(days=1)

    lab_metrics = []
    for lab in sorted(lab_booked_days.keys()):
        booked = len(lab_booked_days[lab])
        pct = booked / total_working_days if total_working_days > 0 else 0
        lab_metrics.append({
            "lab": lab,
            "total_available_days": total_working_days,
            "booked_days": booked,
            "utilization_pct": pct,
        })

    # --- Write to Excel ---
    wb = load_workbook(output_path)
    if "Utilization Metrics" in wb.sheetnames:
        del wb["Utilization Metrics"]
    ws = wb.create_sheet("Utilization Metrics")

    # Section 1: Trainer Utilization
    row = 1
    ws.cell(row=row, column=1, value="Trainer Utilization (Teaching vs Available Days)").font = SECTION_FONT
    row += 1

    trainer_headers = [
        "Trainer", "Location", "Total Working Days", "Leave Days",
        "Available Days", "Teaching Days", "Utilization %",
    ]
    for col_idx, header in enumerate(trainer_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = ALIGN_CENTER
    row += 1

    for tm in trainer_metrics:
        ws.cell(row=row, column=1, value=tm["trainer"])
        ws.cell(row=row, column=2, value=tm["location"]).alignment = ALIGN_CENTER
        ws.cell(row=row, column=3, value=tm["total_working_days"]).alignment = ALIGN_CENTER
        ws.cell(row=row, column=4, value=tm["leave_days"]).alignment = ALIGN_CENTER
        ws.cell(row=row, column=5, value=tm["available_days"]).alignment = ALIGN_CENTER
        ws.cell(row=row, column=6, value=tm["teaching_days"]).alignment = ALIGN_CENTER
        pct_cell = ws.cell(row=row, column=7, value=tm["utilization_pct"])
        pct_cell.number_format = PCT_FORMAT
        pct_cell.alignment = ALIGN_CENTER
        row += 1

    # Section 2: Lab/Room Utilization
    row += 2
    ws.cell(row=row, column=1, value="Lab / Room Utilization (Booked vs Available Days)").font = SECTION_FONT
    row += 1

    lab_headers = ["Lab / Room", "Total Available Days", "Booked Days", "Utilization %"]
    for col_idx, header in enumerate(lab_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = ALIGN_CENTER
    row += 1

    for lm in lab_metrics:
        ws.cell(row=row, column=1, value=lm["lab"])
        ws.cell(row=row, column=2, value=lm["total_available_days"]).alignment = ALIGN_CENTER
        ws.cell(row=row, column=3, value=lm["booked_days"]).alignment = ALIGN_CENTER
        pct_cell = ws.cell(row=row, column=4, value=lm["utilization_pct"])
        pct_cell.number_format = PCT_FORMAT
        pct_cell.alignment = ALIGN_CENTER
        row += 1

    # Column widths
    col_widths = [28, 12, 20, 12, 16, 16, 16]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(output_path)


def plan_courses(session_data: dict) -> dict:
    """Main planning function:
    1. Use pre-computed session requirements
    2. Load trainer availability and priority
    3. Use LLM to assign trainers and dates
    4. Save result to Excel
    """
    try:
        # Step 1: Use session requirements from caller
        courses = session_data["courses"]
        print(f"[PLANNER] Step 1: Got {len(courses)} courses")

        # Step 2: Load trainer data
        print("[PLANNER] Step 2: Loading trainer data...")
        trainers, leave_period = _load_trainer_leave_dates()
        priority_df = _load_trainer_priority()

        # Use the Sep-Dec 2026 planning period
        period = f"{PLANNING_START} to {PLANNING_END}"
        print(f"[PLANNER] Loaded leave dates for {len(trainers)} trainers, planning period: {period}")

        # Step 3: Build prompt context
        print("[PLANNER] Step 3: Building prompt and invoking LLM...")
        priority_records = priority_df.to_dict(orient="records")
        leave_lookup = _build_leave_lookup(trainers)
        trainer_leave_summary = {}
        for trainer_name in _priority_trainer_names(priority_df):
            leave_record = leave_lookup.get(_normalize_trainer_name(trainer_name))
            trainer_leave_summary.setdefault(
                trainer_name,
                {
                    "trainer": trainer_name,
                    "leave_dates": leave_record["leave_dates"] if leave_record else [],
                },
            )
        for trainer in trainers:
            trainer_leave_summary.setdefault(trainer["trainer"], trainer)
        trainer_leave_records = sorted(
            trainer_leave_summary.values(), key=lambda trainer: trainer["trainer"]
        )

        prompt = get_course_planner_prompt()

        llm = get_llm_from_env()
        structured_llm = llm.with_structured_output(CoursePlan)

        chain = prompt | structured_llm

        invoke_input = {
            "period": period,
            "courses_json": json.dumps(courses, indent=2),
            "priority_json": json.dumps(priority_records, indent=2, default=str),
            "trainer_leave_json": json.dumps(
                trainer_leave_records, indent=2
            ),
        }

        # Estimate and print token count
        formatted_messages = prompt.format_messages(**invoke_input)
        full_prompt_text = "\n".join(m.content for m in formatted_messages)
        token_count = _count_tokens(full_prompt_text)
        print(f"[PLANNER] Estimated prompt tokens: {token_count}")
        print(f"[PLANNER] Prompt character count: {len(full_prompt_text)}")

        print("[PLANNER] Invoking LLM (timeout=300s)...")
        start_time = time.time()
        result: CoursePlan = chain.invoke(invoke_input)
        elapsed = time.time() - start_time
        print(f"[PLANNER] LLM returned {len(result.sessions)} sessions in {elapsed:.1f}s")

        # Step 4: Save to Excel
        print("[PLANNER] Step 4: Saving to Excel...")
        rows = [session.model_dump() for session in result.sessions]
        result_df = pd.DataFrame(rows)

        os.makedirs(OUTPUT_DIR, exist_ok=True)
        output_path = os.path.join(OUTPUT_DIR, "course_plan_v3.xlsx")
        result_df.to_excel(output_path, index=False)
        print(f"[PLANNER] Saved to {output_path}")

        # Step 5: Append trainer availability visualization sheet
        try:
            print("[PLANNER] Step 5: Writing trainer availability sheet...")
            _write_trainer_availability_sheet(
                output_path=output_path,
                trainers=trainers,
                sessions=rows,
                priority_trainer_names=_priority_trainer_names(priority_df),
            )
            print("[PLANNER] Trainer availability sheet written")
        except Exception as availability_error:
            print(f"[PLANNER] WARNING: Could not write availability sheet: {availability_error}")
            print(traceback.format_exc())

        # Step 6: Append lab availability sheet
        try:
            print("[PLANNER] Step 6: Writing lab availability sheet...")
            _write_lab_availability_sheet(output_path=output_path, sessions=rows)
            print("[PLANNER] Lab availability sheet written")
        except Exception as lab_error:
            print(f"[PLANNER] WARNING: Could not write lab availability sheet: {lab_error}")
            print(traceback.format_exc())

        # Step 7: Append course overview sheet
        try:
            print("[PLANNER] Step 7: Writing course overview sheet...")
            _write_course_overview_sheet(
                output_path=output_path,
                courses=courses,
                sessions=rows,
            )
            print("[PLANNER] Course overview sheet written")
        except Exception as overview_error:
            print(f"[PLANNER] WARNING: Could not write course overview sheet: {overview_error}")
            print(traceback.format_exc())

        # Step 8: Append utilization metrics sheet
        try:
            print("[PLANNER] Step 8: Writing utilization metrics sheet...")
            _write_utilization_metrics_sheet(
                output_path=output_path,
                trainers=trainers,
                sessions=rows,
                priority_trainer_names=_priority_trainer_names(priority_df),
            )
            print("[PLANNER] Utilization metrics sheet written")
        except Exception as util_error:
            print(f"[PLANNER] WARNING: Could not write utilization metrics sheet: {util_error}")
            print(traceback.format_exc())

        return {
            "total_sessions_planned": len(result.sessions),
            "output_file": "course_plan_v4.xlsx",
            "plan": rows,
        }
    except Exception as e:
        print(f"[PLANNER] ERROR: {e}")
        print(f"[PLANNER] Traceback: {traceback.format_exc()}")
        raise
