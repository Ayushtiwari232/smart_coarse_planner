import json
import os
import tempfile
import time
import traceback
from datetime import date, timedelta

import tiktoken
import pandas as pd
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field

from smart_coarse_planner.Test.prompts import get_course_planner_prompt
from utils.llm_utils import get_llm_from_env


def _count_tokens(text: str, model: str = "gpt-4o") -> int:
    """Estimate token count using tiktoken."""
    try:
        enc = tiktoken.encoding_for_model(model)
    except KeyError:
        enc = tiktoken.get_encoding("cl100k_base")
    return len(enc.encode(text))

INPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "data", "input")
# _DEFAULT_OUTPUT_DIR = os.path.abspath(
#     os.path.join(os.path.dirname(__file__), "..", "data", "output")
# )
# OUTPUT_DIR = os.environ.get("SMART_PLANNER_OUTPUT_DIR") or _DEFAULT_OUTPUT_DIR

_DEFAULT_OUTPUT_DIR = os.path.join(
    tempfile.gettempdir(),
    "smart_planner",
    "output"
)

OUTPUT_DIR = (
    os.environ.get("SMART_PLANNER_OUTPUT_DIR")
    or _DEFAULT_OUTPUT_DIR
)

os.makedirs(OUTPUT_DIR, exist_ok=True)

TRAINER_LEAVE_FILE = "trainer_leave_dates_2026.xlsx"
PRIORITY_FILE = "priority_to_train_list.xlsx"


# --- Pydantic models for structured LLM output ---

class CourseSession(BaseModel):
    course_code: str = Field(description="Course code identifier")
    course_name: str = Field(description="Name of the course")
    session_number: int = Field(description="Session number (1, 2, 3, ...)")
    start_date: str = Field(description="Start date for the session (YYYY-MM-DD)")
    end_date: str = Field(description="End date for the session (YYYY-MM-DD)")
    trainer_name: str = Field(description="Assigned trainer full name")


class CoursePlan(BaseModel):
    sessions: list[CourseSession] = Field(
        description="List of all planned course sessions with dates and trainers"
    )


def _load_trainer_leave_dates() -> tuple[list[dict], str]:
    """Load trainer leave dates from trainer_leave_dates_2026.xlsx."""
    path = os.path.join(INPUT_DIR, TRAINER_LEAVE_FILE)
    df = pd.read_excel(path)

    trainers = []
    all_leave_dates = []
    for _, row in df.iterrows():
        trainer_name = row.get("Trainer Name")
        if pd.isna(trainer_name):
            continue

        raw_leave_dates = row.get("Leave Dates")
        leave_dates = []
        if not pd.isna(raw_leave_dates):
            for raw_date in str(raw_leave_dates).split(","):
                parsed_date = pd.to_datetime(raw_date.strip(), errors="coerce")
                if not pd.isna(parsed_date):
                    leave_dates.append(parsed_date.strftime("%Y-%m-%d"))

        all_leave_dates.extend(leave_dates)
        trainers.append(
            {
                "trainer": str(trainer_name).strip(),
                "leave_dates": sorted(set(leave_dates)),
            }
        )

    years = sorted({pd.to_datetime(date).year for date in all_leave_dates})
    if len(years) == 1:
        period = f"{years[0]}-01-01 to {years[0]}-12-31"
    elif years:
        period = f"{min(all_leave_dates)} to {max(all_leave_dates)}"
    else:
        period = ""

    return trainers, period


def _load_trainer_priority() -> pd.DataFrame:
    """Load trainer priority list. The file has header row in the data,
    so we parse accordingly."""
    path = os.path.join(INPUT_DIR, PRIORITY_FILE)
    df = pd.read_excel(path)
    # First data row is the sub-header: 1, 2, 3, 4, 5
    # Rename columns properly
    df.columns = [
        "course_code",
        "course_name",
        "priority_1",
        "priority_2",
        "priority_3",
        "priority_4",
        "priority_5",
    ]
    # Drop the header row (row 0 which has '1','2','3','4','5')
    df = df.iloc[1:].reset_index(drop=True)
    return df


def _priority_trainer_names(priority_df: pd.DataFrame) -> set[str]:
    priority_columns = [
        "priority_1",
        "priority_2",
        "priority_3",
        "priority_4",
        "priority_5",
    ]
    trainer_names = set()
    for column in priority_columns:
        trainer_names.update(
            str(name).strip()
            for name in priority_df[column].dropna().tolist()
            if str(name).strip()
        )
    return trainer_names


def _normalize_trainer_name(name: str) -> str:
    return " ".join(str(name).strip().lower().split())


def _trainer_name_aliases(name: str) -> set[str]:
    aliases = {_normalize_trainer_name(name)}
    if "," in str(name):
        last_name, first_names = [part.strip() for part in str(name).split(",", 1)]
        aliases.add(_normalize_trainer_name(f"{first_names} {last_name}"))
        last_name_parts = last_name.split()
        if last_name_parts:
            aliases.add(_normalize_trainer_name(f"{first_names} {last_name_parts[0]}"))
    return aliases


def _build_leave_lookup(trainers: list[dict]) -> dict[str, dict]:
    leave_lookup = {}
    for trainer in trainers:
        for alias in _trainer_name_aliases(trainer["trainer"]):
            leave_lookup[alias] = trainer
    return leave_lookup


def _write_trainer_availability_sheet(
    output_path: str,
    trainers: list[dict],
    sessions: list[dict],
    priority_trainer_names: set[str],
) -> None:
    """Append a 'Trainer Availability' sheet to the plan workbook.

    Rows = trainers, Columns = days in the planning period.
    Cell colors:
      - light blue        : trainer is assigned to a course on that day
                            (cell shows course code; hover comment shows
                            course name + session number)
      - light red         : trainer is on leave
      - very light green  : weekend (Sat/Sun)
    """
    LIGHT_BLUE = PatternFill(start_color="FFADD8E6", end_color="FFADD8E6", fill_type="solid")
    LIGHT_RED = PatternFill(start_color="FFFFB6C1", end_color="FFFFB6C1", fill_type="solid")
    VERY_LIGHT_GREEN = PatternFill(start_color="FFE8F5E9", end_color="FFE8F5E9", fill_type="solid")
    HEADER_FILL = PatternFill(start_color="FF1F4E78", end_color="FF1F4E78", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFFFF")
    SESSION_FONT = Font(size=7, bold=True)
    SESSION_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=False)

    # Determine date range from leave dates + session dates
    all_dates: list[date] = []
    for trainer in trainers:
        for raw in trainer.get("leave_dates", []):
            parsed = pd.to_datetime(raw, errors="coerce")
            if not pd.isna(parsed):
                all_dates.append(parsed.date())
    for s in sessions:
        for key in ("start_date", "end_date"):
            parsed = pd.to_datetime(s.get(key), errors="coerce")
            if not pd.isna(parsed):
                all_dates.append(parsed.date())

    if not all_dates:
        print("[PLANNER] Skipping trainer availability sheet: no dates available")
        return

    # Use full calendar year(s) so the visualization always covers Jan–Dec
    years = sorted({d.year for d in all_dates})
    start_d = date(years[0], 1, 1)
    end_d = date(years[-1], 12, 31)

    date_list: list[date] = []
    cursor = start_d
    while cursor <= end_d:
        date_list.append(cursor)
        cursor += timedelta(days=1)

    # Map normalized trainer name -> list of session info dicts
    # Each dict: { start, end, course_code, course_name, session_number }
    trainer_sessions: dict[str, list[dict]] = {}
    for s in sessions:
        norm = _normalize_trainer_name(s.get("trainer_name", ""))
        if not norm:
            continue
        sd = pd.to_datetime(s.get("start_date"), errors="coerce")
        ed = pd.to_datetime(s.get("end_date"), errors="coerce")
        if pd.isna(sd) or pd.isna(ed):
            continue
        trainer_sessions.setdefault(norm, []).append(
            {
                "start": sd.date(),
                "end": ed.date(),
                "course_code": str(s.get("course_code", "") or "").strip(),
                "course_name": str(s.get("course_name", "") or "").strip(),
                "session_number": s.get("session_number"),
            }
        )

    # Combine trainers from the leave file, priority file, and planned sessions
    leave_lookup = _build_leave_lookup(trainers)
    trainer_names: list[str] = []
    seen: set[str] = set()

    def _add_trainer(name: str) -> None:
        cleaned = str(name).strip()
        if not cleaned:
            return
        key = _normalize_trainer_name(cleaned)
        if key in seen:
            return
        seen.add(key)
        trainer_names.append(cleaned)

    for trainer in trainers:
        _add_trainer(trainer["trainer"])
    for name in sorted(priority_trainer_names):
        _add_trainer(name)
    for s in sessions:
        _add_trainer(s.get("trainer_name", ""))

    # Append sheet to the existing workbook
    wb = load_workbook(output_path)
    if "Trainer Availability" in wb.sheetnames:
        del wb["Trainer Availability"]
    ws = wb.create_sheet("Trainer Availability")

    # Header row
    header_cell = ws.cell(row=1, column=1, value="Trainer")
    header_cell.fill = HEADER_FILL
    header_cell.font = HEADER_FONT
    header_cell.alignment = Alignment(horizontal="center", vertical="center")

    for col_idx, dt in enumerate(date_list, start=2):
        cell = ws.cell(row=1, column=col_idx, value=dt.strftime("%Y-%m-%d"))
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(textRotation=90, horizontal="center", vertical="center")

    # Body rows
    for row_idx, trainer_name in enumerate(trainer_names, start=2):
        ws.cell(row=row_idx, column=1, value=trainer_name)

        # Resolve leave dates via all aliases of this name
        leave_set: set[date] = set()
        for alias in _trainer_name_aliases(trainer_name):
            record = leave_lookup.get(alias)
            if not record:
                continue
            for raw in record.get("leave_dates", []):
                parsed = pd.to_datetime(raw, errors="coerce")
                if not pd.isna(parsed):
                    leave_set.add(parsed.date())

        # Resolve session ranges via all aliases
        ranges: list[dict] = []
        for alias in _trainer_name_aliases(trainer_name):
            ranges.extend(trainer_sessions.get(alias, []))

        for col_idx, dt in enumerate(date_list, start=2):
            cell = ws.cell(row=row_idx, column=col_idx)
            matching = [r for r in ranges if r["start"] <= dt <= r["end"]]
            if matching:
                cell.fill = LIGHT_BLUE
                # Show course code(s) inside the cell (short text fits a 4-wide column)
                codes = [r["course_code"] for r in matching if r["course_code"]]
                if codes:
                    cell.value = " / ".join(dict.fromkeys(codes))
                    cell.font = SESSION_FONT
                    cell.alignment = SESSION_ALIGNMENT
                # Hover comment shows full course name + session number
                comment_lines = []
                for r in matching:
                    code = r["course_code"] or "?"
                    name = r["course_name"] or "(unknown course)"
                    session_no = r["session_number"]
                    suffix = f" — Session {session_no}" if session_no else ""
                    comment_lines.append(f"{code}: {name}{suffix}")
                if comment_lines:
                    comment = Comment("\n".join(comment_lines), "Planner")
                    comment.width = 260
                    comment.height = 20 + 14 * len(comment_lines)
                    cell.comment = comment
            elif dt in leave_set:
                cell.fill = LIGHT_RED
            elif dt.weekday() >= 5:
                cell.fill = VERY_LIGHT_GREEN

    # Layout polish
    ws.column_dimensions["A"].width = 28
    for col_idx in range(2, len(date_list) + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 4
    ws.row_dimensions[1].height = 90
    ws.freeze_panes = "B2"

    wb.save(output_path)


def plan_courses(session_data: dict) -> dict:
    """Main planning function:
    1. Use pre-computed session requirements
    2. Load trainer availability and priority
    3. Use LLM to assign trainers and dates
    4. Save result to Excel
    """
    try:
        # Step 1: Use session requirements from caller
        courses = session_data["courses"]
        print(f"[PLANNER] Step 1: Got {len(courses)} courses")

        # Step 2: Load trainer data
        print("[PLANNER] Step 2: Loading trainer data...")
        trainers, period = _load_trainer_leave_dates()
        priority_df = _load_trainer_priority()
        print(f"[PLANNER] Loaded leave dates for {len(trainers)} trainers, period: {period}")

        # Step 3: Build prompt context
        print("[PLANNER] Step 3: Building prompt and invoking LLM...")
        priority_records = priority_df.to_dict(orient="records")
        leave_lookup = _build_leave_lookup(trainers)
        trainer_leave_summary = {}
        for trainer_name in _priority_trainer_names(priority_df):
            leave_record = leave_lookup.get(_normalize_trainer_name(trainer_name))
            trainer_leave_summary.setdefault(
                trainer_name,
                {
                    "trainer": trainer_name,
                    "leave_dates": leave_record["leave_dates"] if leave_record else [],
                },
            )
        for trainer in trainers:
            trainer_leave_summary.setdefault(trainer["trainer"], trainer)
        trainer_leave_records = sorted(
            trainer_leave_summary.values(), key=lambda trainer: trainer["trainer"]
        )

        prompt = get_course_planner_prompt()

        llm = get_llm_from_env()
        structured_llm = llm.with_structured_output(CoursePlan)

        chain = prompt | structured_llm

        invoke_input = {
            "period": period,
            "courses_json": json.dumps(courses, indent=2),
            "priority_json": json.dumps(priority_records, indent=2, default=str),
            "trainer_leave_json": json.dumps(
                trainer_leave_records, indent=2
            ),
        }

        # Estimate and print token count
        formatted_messages = prompt.format_messages(**invoke_input)
        full_prompt_text = "\n".join(m.content for m in formatted_messages)
        token_count = _count_tokens(full_prompt_text)
        print(f"[PLANNER] Estimated prompt tokens: {token_count}")
        print(f"[PLANNER] Prompt character count: {len(full_prompt_text)}")

        print("[PLANNER] Invoking LLM (timeout=300s)...")
        start_time = time.time()
        result: CoursePlan = chain.invoke(invoke_input)
        elapsed = time.time() - start_time
        print(f"[PLANNER] LLM returned {len(result.sessions)} sessions in {elapsed:.1f}s")

        # Step 4: Save to Excel
        print("[PLANNER] Step 4: Saving to Excel...")
        rows = [session.model_dump() for session in result.sessions]
        result_df = pd.DataFrame(rows)

        os.makedirs(OUTPUT_DIR, exist_ok=True)
        output_path = os.path.join(OUTPUT_DIR, "course_plan_v1.xlsx")
        result_df.to_excel(output_path, index=False)
        print(f"[PLANNER] Saved to {output_path}")

        # Step 5: Append trainer availability visualization sheet
        try:
            print("[PLANNER] Step 5: Writing trainer availability sheet...")
            _write_trainer_availability_sheet(
                output_path=output_path,
                trainers=trainers,
                sessions=rows,
                priority_trainer_names=_priority_trainer_names(priority_df),
            )
            print("[PLANNER] Trainer availability sheet written")
        except Exception as availability_error:
            # Don't fail the whole job just because the visualization failed.
            print(f"[PLANNER] WARNING: Could not write availability sheet: {availability_error}")
            print(traceback.format_exc())

        return {
            "total_sessions_planned": len(result.sessions),
            "output_file": "course_plan_v1.xlsx",
            "plan": rows,
        }
    except Exception as e:
        print(f"[PLANNER] ERROR: {e}")
        print(f"[PLANNER] Traceback: {traceback.format_exc()}")
        raise
